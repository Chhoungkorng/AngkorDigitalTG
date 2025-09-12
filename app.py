from flask import Flask, render_template, request, jsonify, send_file, session
from flask_cors import CORS
from telethon import TelegramClient
from telethon.errors import SessionPasswordNeededError, PhoneCodeInvalidError, PhoneNumberInvalidError
from telethon.tl.types import InputPeerUser, InputPeerChannel
import asyncio
import threading
import os
import re
import openpyxl
from io import BytesIO
import time
import json
from datetime import datetime
import logging
from werkzeug.middleware.proxy_fix import ProxyFix
import tempfile
import shutil

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
CORS(app)

# Use environment variable for secret key
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())

# Global variables for client and scraping status
clients = {}  # Store multiple client sessions
scraping_status = {"is_scraping": False, "progress": 0, "status": "Ready", "result": []}
auth_status = {}  # Store authentication status for each session

# Global event loop for async operations
loop = None
loop_thread = None

def get_session_id():
    """Get or create a unique session ID for the user"""
    if 'session_id' not in session:
        session['session_id'] = f"session_{int(time.time())}_{os.urandom(4).hex()}"
    return session['session_id']

def get_session_file(session_id):
    """Get session file path for a specific session - using temp directory for Azure"""
    # Use temp directory that Azure provides
    temp_dir = tempfile.gettempdir()
    sessions_dir = os.path.join(temp_dir, 'telegram_sessions')
    if not os.path.exists(sessions_dir):
        os.makedirs(sessions_dir, exist_ok=True)
    return os.path.join(sessions_dir, f"{session_id}")

def ensure_sessions_dir():
    """Ensure sessions directory exists in temp directory"""
    temp_dir = tempfile.gettempdir()
    sessions_dir = os.path.join(temp_dir, 'telegram_sessions')
    if not os.path.exists(sessions_dir):
        os.makedirs(sessions_dir, exist_ok=True)

def get_or_create_loop():
    """Get or create the global event loop"""
    global loop, loop_thread
    
    if loop is None or loop.is_closed():
        def run_loop():
            global loop
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_forever()
        
        loop_thread = threading.Thread(target=run_loop, daemon=True)
        loop_thread.start()
        
        # Wait for loop to be ready
        while loop is None:
            time.sleep(0.01)
    
    return loop

def run_async(coro):
    """Run an async coroutine in the global event loop"""
    loop = get_or_create_loop()
    future = asyncio.run_coroutine_threadsafe(coro, loop)
    return future.result(timeout=30)  # Add timeout for Azure

async def create_client(api_id, api_hash, session_id):
    """Create a new Telegram client"""
    ensure_sessions_dir()
    session_file = get_session_file(session_id)
    client = TelegramClient(session_file, api_id, api_hash)
    return client

async def send_code(api_id, api_hash, phone, session_id):
    """Send verification code to phone number"""
    try:
        client = await create_client(api_id, api_hash, session_id)
        await client.connect()
        
        result = await client.send_code_request(phone)
        phone_code_hash = result.phone_code_hash
        
        # Store client and hash for verification
        clients[session_id] = {
            'client': client,
            'phone': phone,
            'phone_code_hash': phone_code_hash,
            'api_id': api_id,
            'api_hash': api_hash
        }
        
        return True, "Code sent successfully"
    except PhoneNumberInvalidError:
        return False, "Invalid phone number"
    except Exception as e:
        logger.error(f"Error sending code: {e}")
        return False, f"Error sending code: {str(e)}"

async def verify_code(session_id, code, password=None):
    """Verify the received code and complete authentication"""
    try:
        if session_id not in clients:
            return False, "Session not found. Please request code again."
        
        client_data = clients[session_id]
        client = client_data['client']
        phone = client_data['phone']
        phone_code_hash = client_data['phone_code_hash']
        
        try:
            await client.sign_in(phone, code, phone_code_hash=phone_code_hash)
        except SessionPasswordNeededError:
            if not password:
                return False, "Two-factor authentication enabled. Please provide password."
            await client.sign_in(password=password)
        
        # Test if client is properly authenticated
        me = await client.get_me()
        
        auth_status[session_id] = {
            'authenticated': True,
            'user_id': me.id,
            'phone': me.phone,
            'username': me.username,
            'first_name': me.first_name,
            'last_name': me.last_name,
            'login_time': datetime.now().isoformat()
        }
        
        return True, f"Successfully logged in as {me.first_name or me.username}"
        
    except PhoneCodeInvalidError:
        return False, "Invalid verification code"
    except Exception as e:
        logger.error(f"Authentication error: {e}")
        return False, f"Authentication error: {str(e)}"

async def logout_client(session_id):
    """Logout and cleanup client session"""
    try:
        if session_id in clients:
            client = clients[session_id]['client']
            if client.is_connected():
                await client.log_out()
            await client.disconnect()
            del clients[session_id]
        
        if session_id in auth_status:
            del auth_status[session_id]
        
        # Remove session file
        session_file = get_session_file(session_id)
        if os.path.exists(f"{session_file}.session"):
            try:
                os.remove(f"{session_file}.session")
            except:
                pass  # Ignore cleanup errors
        
        return True, "Logged out successfully"
    except Exception as e:
        logger.error(f"Logout error: {e}")
        return False, f"Logout error: {str(e)}"

def is_authenticated(session_id):
    """Check if user is authenticated"""
    return session_id in auth_status and auth_status[session_id].get('authenticated', False)

async def scrape_data(group, scrape_type, session_id):
    """Scrape data from Telegram group"""
    global scraping_status
    try:
        if not is_authenticated(session_id):
            raise Exception("Not authenticated. Please login first.")
        
        client = clients[session_id]['client']
        
        if not client.is_connected():
            await client.connect()
        
        # Update status
        scraping_status["status"] = "Connecting to group..."
        scraping_status["progress"] = 20
        
        # Get entity from group link/username
        entity = await client.get_entity(group)
        
        # Update status
        scraping_status["status"] = "Fetching participants..."
        scraping_status["progress"] = 40
        
        # Get participants with a reasonable limit for Azure timeout
        participants = await client.get_participants(entity, limit=5000)
        
        # Update status
        scraping_status["status"] = "Processing data..."
        scraping_status["progress"] = 70
        
        scraped_data = []
        for user in participants:
            if scrape_type == "Phone Numbers" and user.phone:
                scraped_data.append(f"+{user.phone}")
            elif scrape_type == "Usernames" and user.username:
                scraped_data.append(f"@{user.username}")
            elif scrape_type == "User IDs":
                scraped_data.append(f"https://web.telegram.org/k/#{user.id}")
            elif scrape_type == "All Data":
                user_data = {
                    "id": user.id,
                    "username": f"@{user.username}" if user.username else None,
                    "phone": f"+{user.phone}" if user.phone else None,
                    "first_name": user.first_name,
                    "last_name": user.last_name
                }
                scraped_data.append(user_data)
        
        # Update status
        scraping_status["status"] = f"Successfully scraped {len(scraped_data)} items"
        scraping_status["progress"] = 100
        scraping_status["result"] = scraped_data
        
        return scraped_data
        
    except Exception as e:
        error_msg = f"Error: {str(e)}"
        scraping_status["status"] = error_msg
        scraping_status["progress"] = 0
        logger.error(f"Scraping error: {e}")
        raise e

def run_scraping_task(group, scrape_type, session_id):
    """Run scraping task in the background"""
    global scraping_status
    try:
        result = run_async(scrape_data(group, scrape_type, session_id))
        return result
    except Exception as e:
        scraping_status["status"] = f"Error: {str(e)}"
        scraping_status["progress"] = 0
        scraping_status["is_scraping"] = False
        logger.error(f"Scraping task error: {e}")

@app.route('/')
def index():
    session_id = get_session_id()
    return render_template('index.html')

@app.route('/api/auth/send_code', methods=['POST'])
def api_send_code():
    """API endpoint to send verification code"""
    data = request.json
    api_id = data.get('api_id')
    api_hash = data.get('api_hash')
    phone = data.get('phone')
    
    if not all([api_id, api_hash, phone]):
        return jsonify({"error": "API ID, API Hash, and phone number are required"}), 400
    
    try:
        api_id = int(api_id)
    except ValueError:
        return jsonify({"error": "API ID must be a number"}), 400
    
    session_id = get_session_id()
    
    try:
        success, message = run_async(send_code(api_id, api_hash, phone, session_id))
        if success:
            return jsonify({"message": message, "session_id": session_id})
        else:
            return jsonify({"error": message}), 400
    except Exception as e:
        logger.error(f"Send code API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/auth/verify_code', methods=['POST'])
def api_verify_code():
    """API endpoint to verify code and complete authentication"""
    data = request.json
    code = data.get('code')
    password = data.get('password')
    
    if not code:
        return jsonify({"error": "Verification code is required"}), 400
    
    session_id = get_session_id()
    
    try:
        success, message = run_async(verify_code(session_id, code, password))
        if success:
            return jsonify({"message": message, "user_info": auth_status.get(session_id, {})})
        else:
            return jsonify({"error": message}), 400
    except Exception as e:
        logger.error(f"Verify code API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/auth/status')
def api_auth_status():
    """API endpoint to get authentication status"""
    session_id = get_session_id()
    authenticated = is_authenticated(session_id)
    
    if authenticated:
        return jsonify({
            "authenticated": True,
            "user_info": auth_status[session_id]
        })
    else:
        return jsonify({"authenticated": False})

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    """API endpoint to logout"""
    session_id = get_session_id()
    
    try:
        success, message = run_async(logout_client(session_id))
        # Clear Flask session
        session.clear()
        
        if success:
            return jsonify({"message": message})
        else:
            return jsonify({"error": message}), 400
    except Exception as e:
        logger.error(f"Logout API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/scrape', methods=['POST'])
def api_scrape():
    """API endpoint to start scraping"""
    global scraping_status
    
    session_id = get_session_id()
    
    if not is_authenticated(session_id):
        return jsonify({"error": "Please login to Telegram first"}), 401
    
    if scraping_status["is_scraping"]:
        return jsonify({"error": "Scraping is already in progress"}), 400
    
    data = request.json
    group = data.get('group')
    scrape_type = data.get('scrape_type')
    
    if not group:
        return jsonify({"error": "Group link is required"}), 400
    
    # Start scraping in a separate thread
    scraping_status["is_scraping"] = True
    scraping_status["progress"] = 10
    scraping_status["status"] = "Initializing..."
    scraping_status["result"] = []
    
    def scraping_thread():
        try:
            run_scraping_task(group, scrape_type, session_id)
        finally:
            scraping_status["is_scraping"] = False
    
    thread = threading.Thread(target=scraping_thread, daemon=True)
    thread.start()
    
    return jsonify({"message": "Scraping started"})

@app.route('/api/status')
def api_status():
    """API endpoint to get scraping status"""
    return jsonify(scraping_status)

@app.route('/api/export', methods=['POST'])
def api_export():
    """API endpoint to export data"""
    session_id = get_session_id()
    
    if not is_authenticated(session_id):
        return jsonify({"error": "Please login to Telegram first"}), 401
    
    data = request.json
    export_data = data.get('data', [])
    format_type = data.get('format', 'txt')
    filename = data.get('filename', 'scraped_data')
    
    if not export_data:
        return jsonify({"error": "No data to export"}), 400
    
    try:
        if format_type == 'txt':
            # Create text file
            if isinstance(export_data[0], dict):
                content = "\n".join([json.dumps(item, ensure_ascii=False) for item in export_data])
            else:
                content = "\n".join(export_data)
            return jsonify({"content": content, "filename": f"{filename}.txt"})
        
        elif format_type == 'xlsx':
            # Create Excel file in memory
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scraped Data"
            
            if isinstance(export_data[0], dict):
                # Write headers for All Data
                headers = ["ID", "Username", "Phone", "First Name", "Last Name"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col).value = header
                
                # Write data
                for row, item in enumerate(export_data, 2):
                    ws.cell(row=row, column=1).value = item.get("id")
                    ws.cell(row=row, column=2).value = item.get("username")
                    ws.cell(row=row, column=3).value = item.get("phone")
                    ws.cell(row=row, column=4).value = item.get("first_name")
                    ws.cell(row=row, column=5).value = item.get("last_name")
            else:
                ws['A1'] = "Scraped Data"
                for i, item in enumerate(export_data, 2):
                    ws[f'A{i}'] = item
            
            # Save to bytes buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return send_file(
                buffer,
                as_attachment=True,
                download_name=f"{filename}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        else:
            return jsonify({"error": "Unsupported format"}), 400
            
    except Exception as e:
        logger.error(f"Export error: {e}")
        return jsonify({"error": f"Export failed: {str(e)}"}), 500

@app.route('/api/process', methods=['POST'])
def api_process():
    """API endpoint to process data (remove duplicates, filter, etc.)"""
    session_id = get_session_id()
    
    if not is_authenticated(session_id):
        return jsonify({"error": "Please login to Telegram first"}), 401
    
    data = request.json
    process_type = data.get('type')
    input_data = data.get('data', [])
    param = data.get('param', '')
    
    if not input_data:
        return jsonify({"error": "No data to process"}), 400
    
    try:
        if process_type == "remove_duplicates":
            # Remove duplicates while preserving order
            if isinstance(input_data[0], dict):
                # For dictionaries, use ID as key for uniqueness
                seen = set()
                result = []
                for item in input_data:
                    if item["id"] not in seen:
                        seen.add(item["id"])
                        result.append(item)
            else:
                result = list(dict.fromkeys(input_data))
            return jsonify({"result": result, "removed": len(input_data) - len(result)})
        
        elif process_type == "remove_word":
            # Remove lines containing specific word
            if not param:
                return jsonify({"error": "Word parameter is required"}), 400
            
            if isinstance(input_data[0], dict):
                result = [item for item in input_data if not any(param.lower() in str(value).lower() for value in item.values() if value)]
            else:
                result = [line for line in input_data if param.lower() not in line.lower()]
            return jsonify({"result": result, "removed": len(input_data) - len(result)})
        
        elif process_type == "filter_sim":
            # Filter by SIM company
            if not param:
                return jsonify({"error": "SIM company parameter is required"}), 400
            
            # Define prefixes for each company (Cambodia)
            prefixes = {
                "Smart": ["85510", "85515", "85516", "85569", "85570", "85581", "85586", "85587", "85593", "85596", "85598"],
                "Cellcard": ["85511", "85512", "85514", "85517", "85561", "85576", "85577", "85578", "85579", "85585", "85589", "85592", "85595", "85599"],
                "Metfone": ["85588", "85597", "85571", "85560", "85566", "85567", "85568", "85590"]
            }
            
            selected_prefixes = prefixes.get(param, [])
            result = []
            
            for item in input_data:
                if isinstance(item, dict):
                    phone = item.get("phone")
                    if not phone:
                        continue
                    numbers = re.findall(r'(\+?855|0)?(\d{8,})', phone)
                else:
                    numbers = re.findall(r'(\+?855|0)?(\d{8,})', item)
                
                for number in numbers:
                    full_number = number[1]
                    if len(full_number) >= 8:
                        normalized = f"855{full_number[-8:]}" if len(full_number) == 8 else f"855{full_number[-9:]}"
                        for prefix in selected_prefixes:
                            if normalized.startswith(prefix):
                                result.append(item)
                                break
                        else:
                            continue
                        break
            
            return jsonify({"result": result, "removed": len(input_data) - len(result)})
        
        elif process_type == "filter_855":
            # Filter 855 numbers
            result = []
            for item in input_data:
                if isinstance(item, dict):
                    phone = item.get("phone")
                    if phone and re.search(r'(\+?855)\d{8,9}', phone):
                        result.append(item)
                elif re.search(r'(\+?855)\d{8,9}', item):
                    result.append(item)
            
            return jsonify({"result": result, "removed": len(input_data) - len(result)})
        
        elif process_type == "find_facebook":
            # Find Facebook URLs
            result = []
            fb_patterns = [
                r"https?://(www\.)?facebook\.com/profile\.php\?id=\d+",
                r"https?://(www\.)?facebook\.com/[a-zA-Z0-9._-]+",
                r"https?://(m\.)?facebook\.com/[a-zA-Z0-9._-]+",
                r"https?://fb\.com/[a-zA-Z0-9._-]+",
            ]
            
            for item in input_data:
                text = json.dumps(item, ensure_ascii=False) if isinstance(item, dict) else item
                for pattern in fb_patterns:
                    if re.search(pattern, text, re.IGNORECASE):
                        result.append(item)
                        break
            
            return jsonify({"result": result, "removed": len(input_data) - len(result)})
        
        elif process_type == "filter_fb_id":
            # Filter Facebook ID URLs only
            result = []
            pattern = r"https?://(www\.|m\.)?facebook\.com/profile\.php\?id=\d+"
            
            for item in input_data:
                text = json.dumps(item, ensure_ascii=False) if isinstance(item, dict) else item
                if re.search(pattern, text, re.IGNORECASE):
                    result.append(item)
            
            return jsonify({"result": result, "removed": len(input_data) - len(result)})
        
        elif process_type == "extract_ids":
            # Extract Telegram user IDs
            result = []
            for item in input_data:
                if isinstance(item, dict):
                    if item.get("id"):
                        result.append(str(item["id"]))
                elif isinstance(item, str) and item.startswith("https://web.telegram.org/k/#"):
                    try:
                        user_id = item.split("#")[-1]
                        if user_id.isdigit():
                            result.append(user_id)
                    except:
                        continue
            return jsonify({"result": result, "removed": len(input_data) - len(result)})
        
        elif process_type == "convert_to_web":
            # Convert user IDs to Telegram web links
            result = []
            for item in input_data:
                if isinstance(item, dict):
                    if item.get("id"):
                        result.append(f"https://web.telegram.org/k/#{item['id']}")
                elif isinstance(item, str) and item.isdigit():
                    result.append(f"https://web.telegram.org/k/#{item}")
            return jsonify({"result": result, "removed": len(input_data) - len(result)})
        
        else:
            return jsonify({"error": "Unknown process type"}), 400
            
    except Exception as e:
        logger.error(f"Processing error: {e}")
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500

# Cleanup function to stop the event loop when the app shuts down
def cleanup():
    global loop, loop_thread
    if loop and not loop.is_closed():
        loop.call_soon_threadsafe(loop.stop)
    if loop_thread and loop_thread.is_alive():
        loop_thread.join(timeout=5)

import atexit
atexit.register(cleanup)

if __name__ == '__main__':
    ensure_sessions_dir()
    # Use environment variable for port, defaulting to 80 for Azure
    port = int(os.environ.get('PORT', 80))
    app.run(host='0.0.0.0', port=port, debug=False)